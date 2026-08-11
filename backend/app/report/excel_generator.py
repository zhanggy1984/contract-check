"""Excel 校验报告生成（T4.4）：openpyxl，三个 sheet（基本信息/校验明细/异常明细）。"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.report.report_data import ReportData

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
LABEL_FONT = Font(name="微软雅黑", size=10, bold=True, color="595959")
WRAP = Alignment(vertical="top", wrap_text=True)
EV_FILL = PatternFill("solid", fgColor="FFF7E6")


def _write_kv(ws, rows: list[tuple[str, str]], start_row: int) -> int:
    for i, (k, v) in enumerate(rows):
        ws.cell(row=start_row + i, column=1, value=k).font = LABEL_FONT
        ws.cell(row=start_row + i, column=2, value=v).font = BODY_FONT
        ws.cell(row=start_row + i, column=2).alignment = WRAP
    return start_row + len(rows)


def _sheet_title(ws, text: str) -> None:
    ws.merge_cells("A1:C1")
    ws.cell(row=1, column=1, value=text).font = TITLE_FONT
    ws.row_dimensions[1].height = 24


def render(data: ReportData) -> BytesIO:
    wb = Workbook()

    # ---- Sheet1 基本信息与抽取摘要 ----
    ws = wb.active
    ws.title = "基本信息"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 70
    _sheet_title(ws, f"合同校验报告（任务号 {data.task_id}）")
    row = 3
    base = [
        ("合同文件", data.file_name), ("文件类型", data.file_type), ("文件大小", data.file_size),
        ("是否扫描件", "是（已 OCR）" if data.has_scanned and data.ocr_applied
         else "是（未 OCR）" if data.has_scanned else "否（含文本层）"),
        ("任务状态", data.task_status), ("抽取状态", data.extraction_status or ""),
        ("LLM 模型", data.llm_model or ""), ("创建时间", data.create_time or ""),
    ]
    row = _write_kv(ws, base, row)
    if data.summary:
        row += 1
        ws.cell(row=row, column=1, value="抽取摘要").font = TITLE_FONT
        row += 1
        row = _write_kv(ws, data.summary, row)
    for title, entities in (("当事人", data.parties), ("标的明细", data.items)):
        for i, fields in enumerate(entities, start=1):
            row += 1
            ws.cell(row=row, column=1, value=f"{title} {i}").font = LABEL_FONT
            row += 1
            row = _write_kv(ws, fields, row)

    # ---- Sheet2 校验明细 ----
    ws2 = wb.create_sheet("校验明细")
    headers = ["规则名称", "规则 ID", "类型", "结果", "严重级别", "置信度", "段落", "说明"]
    for c, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
    widths = [28, 10, 10, 8, 9, 9, 12, 60]
    for c, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(c)].width = w
    for i, r in enumerate(data.rule_results, start=2):
        vals = [r["rule_name"], r["rule_id"], r["rule_type"], r["result"], r["severity"],
                r["confidence"], r["segment_ref"], r["message"]]
        for c, v in enumerate(vals, start=1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.font = BODY_FONT
            cell.alignment = WRAP
    ws2.freeze_panes = "A2"

    # ---- Sheet3 异常明细 ----
    ws3 = wb.create_sheet("异常明细")
    headers3 = ["规则名称", "规则 ID", "类型", "严重级别", "置信度", "段落", "审核状态",
                "确认人", "确认时间", "问题说明", "原文证据"]
    for c, h in enumerate(headers3, start=1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
    widths3 = [26, 10, 10, 9, 9, 12, 9, 9, 18, 45, 55]
    for c, w in enumerate(widths3, start=1):
        ws3.column_dimensions[get_column_letter(c)].width = w
    for i, v in enumerate(data.violations, start=2):
        vals = [v["rule_name"], v["rule_id"], v["rule_type"], v["severity"], v["confidence"],
                v["segment_ref"], v["status"], v["confirm_user"], v["confirm_time"],
                v["message"], v["evidence_text"]]
        for c, val in enumerate(vals, start=1):
            cell = ws3.cell(row=i, column=c, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            if c == 11 and val:
                cell.fill = EV_FILL
    ws3.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
