"""报告导出单元测试（T4.4）：report_data 组装逻辑 + PDF/Excel 生成冒烟。

build_report_data 依赖 DB，用 SimpleNamespace 假对象隔离（只读属性，不触真实 ORM）；
PDF 渲染依赖中文字体，无字体环境跳过（不硬失败）。
用 unittest 风格（与 test_semantic_evaluator.py 一致），pytest 仅作 runner。
"""
import unittest
from datetime import datetime
from types import SimpleNamespace

from app.report import excel_generator, pdf_generator
from app.report.font import ensure_registered
from app.report.report_data import ReportData, build_report_data

STD_JSON = """{
  "contractTitle": "设备采购与服务合同",
  "contractType": "采购",
  "totalAmount": 500000,
  "hasParty": [{"partyRole": "甲方", "partyName": "北京某科技有限公司"}],
  "hasItem": [{"itemName": "智能巡检设备", "itemAmount": 500000}]
}"""


def _task() -> SimpleNamespace:
    return SimpleNamespace(
        id=30, status="WAITING_REVIEW", extraction_status="COMPLETE",
        llm_model="deepseek-chat", create_time=datetime(2026, 8, 11, 8, 0, 0),
        contract_file=SimpleNamespace(
            file_name="scanned_test.pdf", file_type="PDF", file_size=11330688,
            has_scanned=True, ocr_applied=True),
        standard_json=STD_JSON,
    )


def _rule(id_, rule_type="DETERMINISTIC"):
    return SimpleNamespace(id=id_, rule_name=f"规则{id_}", rule_type=rule_type)


def _rcr(rule_id, result="PASS"):
    return SimpleNamespace(
        rule_id=rule_id, rule_type="DETERMINISTIC", result=result, severity="HIGH",
        confidence="HIGH", segment_ref=None, message="",
    )


def _viol(rule_id):
    return SimpleNamespace(
        rule_id=rule_id, rule_type="SEMANTIC", severity="HIGH", confidence="HIGH",
        segment_ref="seg-4", evidence_text="任何一方违约，应支付违约金",
        message="违约条款不完整", status="UNCONFIRMED", confirm_user=None,
        confirm_time=None, create_time=datetime(2026, 8, 11, 8, 1, 0),
    )


class _Q:
    """模拟 db.query(...).filter_by(...).order_by(...).all() 链式调用。"""

    def __init__(self, rows):
        self._rows = rows

    def filter_by(self, **kwargs):
        return self

    def order_by(self, *args, **kwargs):
        return self

    def all(self):
        return self._rows


class _FakeDb:
    def __init__(self, task, rules, rcr, viol):
        self.task = task
        self._queries = {type(r).__name__: _Q(rules) if r is rules else None for r in []}
        self._rules = _Q(rules)
        self._rcr = _Q(rcr)
        self._viol = _Q(viol)

    def get(self, model, task_id):
        return self.task if task_id == self.task.id else None

    def query(self, model):
        name = model.__name__
        return {"CheckRule": self._rules, "RuleCheckResult": self._rcr,
                "Violation": self._viol}.get(name, _Q([]))


class TestBuildReportData(unittest.TestCase):
    def _data(self):
        db = _FakeDb(_task(), [_rule(1), _rule(26, "SEMANTIC")], [_rcr(1), _rcr(26, "FAIL")], [_viol(26)])
        return build_report_data(db, 30)

    def test_missing_task_raises(self):
        db = _FakeDb(_task(), [], [], [])
        with self.assertRaises(ValueError):
            build_report_data(db, 999)

    def test_summary_label_mapping(self):
        d = self._data()
        kv = dict(d.summary)
        # 中文标签而非原始 JSON 键；空值字段不出现
        self.assertEqual(kv.get("合同名称"), "设备采购与服务合同")
        self.assertNotIn("contractType", kv)
        self.assertNotIn("taxRate", kv)  # 缺失字段不展示

    def test_parties_and_items(self):
        d = self._data()
        self.assertEqual(len(d.parties), 1)
        self.assertEqual(dict(d.parties[0]).get("名称"), "北京某科技有限公司")
        self.assertEqual(len(d.items), 1)
        self.assertEqual(dict(d.items[0]).get("标的名称"), "智能巡检设备")

    def test_rule_results_carries_rule_name(self):
        d = self._data()
        self.assertEqual(len(d.rule_results), 2)
        by_rule = {r["rule_id"]: r for r in d.rule_results}
        self.assertEqual(by_rule[26]["rule_name"], "规则26")

    def test_violation_carries_evidence_and_label(self):
        d = self._data()
        v = d.violations[0]
        self.assertEqual(v["evidence_text"], "任何一方违约，应支付违约金")
        self.assertEqual(v["status"], "待确认")  # UNCONFIRMED → 中文标签

    def test_scanned_file_flag(self):
        d = self._data()
        self.assertTrue(d.has_scanned and d.ocr_applied)


class TestGenerators(unittest.TestCase):
    def _sample(self):
        return ReportData(
            task_id=1, task_status="SUCCESS", extraction_status="COMPLETE",
            llm_model="deepseek-chat", create_time="2026-08-11T00:00:00",
            file_name="测试合同.pdf", file_type="PDF", file_size="1.0 KB",
            has_scanned=False, ocr_applied=False,
            summary=[("合同名称", "测试合同")],
            parties=[[("角色", "甲方"), ("名称", "甲公司")]],
            items=[[("标的名称", "服务")]],
            rule_results=[{"rule_id": 1, "rule_name": "规则一", "rule_type": "SEMANTIC",
                           "result": "FAIL", "severity": "HIGH", "confidence": "HIGH",
                           "segment_ref": "seg-0", "message": "缺违约条款"}],
            violations=[{"rule_id": 1, "rule_name": "规则一", "rule_type": "SEMANTIC",
                         "severity": "HIGH", "confidence": "HIGH", "segment_ref": "seg-0",
                         "evidence_text": "任何一方违约，应支付违约金",
                         "message": "缺违约条款", "status": "待确认",
                         "confirm_user": "", "confirm_time": ""}],
        )

    def test_pdf_generates_valid_header(self):
        try:
            ensure_registered()
        except RuntimeError:
            self.skipTest("无中文字体，跳过 PDF 渲染测试")
        buf = pdf_generator.render(self._sample())
        self.assertTrue(buf.getvalue().startswith(b"%PDF"))

    def test_excel_generates_three_sheets(self):
        from openpyxl import load_workbook
        from io import BytesIO
        buf = excel_generator.render(self._sample())
        wb = load_workbook(BytesIO(buf.getvalue()))
        self.assertEqual(wb.sheetnames, ["基本信息", "校验明细", "异常明细"])


if __name__ == "__main__":
    unittest.main()
